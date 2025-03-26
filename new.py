import os
import cv2
import datetime
import geocoder
import numpy as np
from PIL import Image
import mediapipe as mp
from transformers import CLIPProcessor, CLIPModel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load CLIP model
model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32")
processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")

# Create or load Excel workbook
excel_file = "face_attendance.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Timestamp", "Name", "Location", "Confidence"])
    wb.save(excel_file)

# Location once per session
def get_location():
    try:
        g = geocoder.ip("me")
        if g.latlng:
            return f"{g.latlng[0]},{g.latlng[1]}"
        return "Not Found"
    except:
        return "Error"

location = get_location()

# Load known face embeddings
def get_clip_embedding(pil_image):
    inputs = processor(images=pil_image, return_tensors="pt")
    with torch.no_grad():
        outputs = model.get_image_features(**inputs)
    return outputs[0].numpy()

known_faces = {}
os.makedirs("known_faces", exist_ok=True)
for file in os.listdir("known_faces"):
    if file.endswith((".jpg", ".png")):
        name = os.path.splitext(file)[0]
        image = Image.open(os.path.join("known_faces", file)).convert("RGB")
        embedding = get_clip_embedding(image)
        known_faces[name] = embedding

# Cosine similarity
def cosine_similarity(a, b):
    return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))

# Face detection
mp_face = mp.solutions.face_detection
face_detection = mp_face.FaceDetection(model_selection=0, min_detection_confidence=0.6)

# Start webcam
cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = face_detection.process(rgb_frame)

    if results.detections:
        for detection in results.detections:
            bboxC = detection.location_data.relative_bounding_box
            h, w, _ = frame.shape
            x, y, bw, bh = int(bboxC.xmin * w), int(bboxC.ymin * h), int(bboxC.width * w), int(bboxC.height * h)
            face_crop = rgb_frame[y:y+bh, x:x+bw]

            try:
                face_img = Image.fromarray(face_crop).resize((224, 224)).convert("RGB")
                face_embedding = get_clip_embedding(face_img)

                name = "Unknown"
                max_sim = 0.0
                for known_name, known_emb in known_faces.items():
                    sim = cosine_similarity(face_embedding, known_emb)
                    if sim > 0.80 and sim > max_sim:
                        name = known_name
                        max_sim = sim

                # Draw box
                cv2.rectangle(frame, (x, y), (x+bw, y+bh), (0, 255, 0), 2)
                cv2.putText(frame, f"{name} ({max_sim:.2f})", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)

                # Log to Excel
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                wb = load_workbook(excel_file)
                ws = wb.active
                ws.append([timestamp, name, location, f"{max_sim:.2f}"])
                wb.save(excel_file)

            except Exception as e:
                print("Face crop error:", e)

    cv2.imshow("Face Recognition with Excel Log", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
